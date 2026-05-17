"""Parse taxa rows from the external plant knowledge-base workbook."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


PLANTS_SHEET_NAME = "plants"

RANKS = (
    ("ordo", "目", "Ordo"),
    ("familia", "科", "Familia"),
    ("genus", "属", "Genus"),
    ("species", "种", "Species"),
)
BINOMIAL_KEY = "binomial_name"


@dataclass(frozen=True)
class TaxaRow:
    ordo: str | None = None
    familia: str | None = None
    genus: str | None = None
    species: str | None = None
    binomial_name: str | None = None


def read_taxa_rows(
    workbook_path: str | Path,
    max_rows: int | None = None,
) -> Iterable[TaxaRow]:
    """Yield selected taxa columns from the workbook's plants sheet."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to import taxa workbooks") from exc

    workbook = load_workbook(
        Path(workbook_path).expanduser(),
        read_only=True,
        data_only=True,
        keep_vba=False,
    )
    worksheet = workbook[PLANTS_SHEET_NAME]
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows)
    columns = detect_columns(headers)

    for offset, row in enumerate(rows, start=2):
        if max_rows is not None and offset > max_rows:
            break
        parsed = TaxaRow(
            ordo=_clean_cell(row[columns["ordo"]]),
            familia=_clean_cell(row[columns["familia"]]),
            genus=_clean_cell(row[columns["genus"]]),
            species=_clean_cell(row[columns["species"]]),
            binomial_name=_clean_cell(row[columns[BINOMIAL_KEY]]),
        )
        if any((parsed.ordo, parsed.familia, parsed.genus, parsed.species)):
            yield parsed


def detect_columns(headers: tuple[object, ...]) -> dict[str, int]:
    """Find target columns even when the source header has extra characters."""

    columns: dict[str, int] = {}
    normalized = [_normalize_header(header) for header in headers]

    for key, chinese, latin in RANKS:
        columns[key] = _find_header(normalized, chinese, latin)
    columns[BINOMIAL_KEY] = _find_header(normalized, "学名", "Binomial name")
    return columns


def _find_header(headers: list[str], chinese: str, latin: str) -> int:
    for index, header in enumerate(headers):
        if header.startswith(chinese) and latin.lower() in header.lower():
            return index
    raise ValueError(f"Missing required taxa column: {chinese}({latin})")


def _normalize_header(value: object) -> str:
    return str(value or "").strip()


def _clean_cell(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
